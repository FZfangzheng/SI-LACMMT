from PIL import Image
import numpy as np
import os
from skimage import measure
import openpyxl
IMG_EXTENSIONS = [
    '.jpg', '.JPG', '.jpeg', '.JPEG',
    '.png', '.PNG', '.ppm', '.PPM', '.bmp', '.BMP','.tif','.npy'
]


def is_image_file(filename): # npy文件现在也算
    return any(filename.endswith(extension) for extension in IMG_EXTENSIONS)


def make_dataset(dir):
    images = []
    assert os.path.isdir(dir), '%s is not a valid directory' % dir

    for root, _, fnames in sorted(os.walk(dir)):
        for fname in fnames:
            if is_image_file(fname):
                path = os.path.join(root, fname)
                images.append(path)
    return images
def get_inner_path(file_path,floder_path):
    assert file_path[:len(floder_path)]==floder_path,"传入的文件不在文件夹中！[%s][%s]"%(file_path,floder_path)
    file_path=file_path[len(floder_path)+1:]
    return file_path

def rgb2gray(rgb):
    return np.dot(rgb[..., :3], [0.299, 0.587, 0.114]).astype(np.uint8)

def scores_from_floder(path1,path2):
    ssim = 0
    ssim_rgb=0
    mse=0
    psnr=0
    num = 0
    imgs1 = make_dataset(path1)
    imgs2 = make_dataset(path2)
    from tqdm import tqdm
    for img1 in tqdm(imgs1):
        img_inner = get_inner_path(img1, path1)
        if os.path.join(path2, img_inner) in imgs2:
            img2 = os.path.join(path2, img_inner)
            img1_PIL = Image.open(img1)
            img2_PIL = Image.open(img2)
            img1_np = np.array(img1_PIL)
            img2_np = np.array(img2_PIL)
            # RGB to 灰度
            # if len(img1_np.shape) > 2:
            #     img1_np = rgb2gray(img1_np)
            # if len(img2_np.shape) > 2:
            #     img2_np = rgb2gray(img2_np)
            ssim_tmp = measure.compare_ssim(rgb2gray(img1_np),rgb2gray( img2_np))
            ssim_rgb_tmp=measure.compare_ssim(img1_np,img2_np,multichannel=True)
            mse_tmp=measure.compare_mse(img1_np,img2_np)
            psnr_tmp=measure.compare_psnr(img1_np,img2_np)
            ssim += ssim_tmp
            ssim_rgb+=ssim_rgb_tmp
            mse+=mse_tmp
            psnr+=psnr_tmp
            num += 1
    assert num > 0
    ssim = ssim / num
    ssim_rgb=ssim_rgb/num
    mse=mse/num
    psnr=psnr/num
    return ssim,mse, psnr,ssim_rgb


if __name__=='__main__':
    # real_path=r'D:\map_translate\看看效果\0605p2p数据S2OMGAN默认参数，epoch200\real_result'
    # fake_path=r'D:\map_translate\看看效果\0605p2p数据S2OMGAN默认参数，epoch200\fake_result'

    wb = openpyxl.Workbook()
    wb.create_sheet('origin_datas')
    sh = wb['origin_datas']
    sh.cell(row=1, column=2, value='mse')
    sh.cell(row=1, column=3, value='ssim')
    sh.cell(row=1, column=4, value='psnr')
    sh.cell(row=1, column=5, value='ssim_rgb')
    #
    # real_paths = [r'D:\map_translate\看看效果\0704重绘方式1的补充试验\数据集对比试验\p2pHD模型-在p2p数据集上训练在TW16上生成\real_result',
    #               r'D:\map_translate\看看效果\0704重绘方式1的补充试验\数据集对比试验\p2pHD模型-在TW16上训练&生成\real_result',
    #               r'D:\map_translate\看看效果\0704重绘方式1的补充试验\数据集对比试验\结合模型-在p2p数据集上训练在TW16上生成\real_result',
    #               r'D:\map_translate\看看效果\0704重绘方式1的补充试验\数据集对比试验\结合模型-在TW16上训练&生成\real_result']
    #
    # fake_paths = [r'D:\map_translate\看看效果\0704重绘方式1的补充试验\数据集对比试验\p2pHD模型-在p2p数据集上训练在TW16上生成\fake_result',
    #               r'D:\map_translate\看看效果\0704重绘方式1的补充试验\数据集对比试验\p2pHD模型-在TW16上训练&生成\fake_result',
    #               r'D:\map_translate\看看效果\0704重绘方式1的补充试验\数据集对比试验\结合模型-在p2p数据集上训练在TW16上生成\fake_result',
    #               r'D:\map_translate\看看效果\0704重绘方式1的补充试验\数据集对比试验\结合模型-在TW16上训练&生成\fake_result']

    real_paths = [r'D:\map_translate\看看效果\0627两个重绘数据集试验\color1-结合方法\real_result']
                  # r'D:\map_translate\看看效果\0627两个重绘数据集试验\color1-结合方法\real_result',
                  # r'D:\map_translate\看看效果\0627两个重绘数据集试验\color1-纯p2pHD\real_result']

    fake_paths = [r'D:\map_translate\看看效果\0627两个重绘数据集试验\color1-结合方法\fake_result_tostand']
                  # r'D:\map_translate\看看效果\0627两个重绘数据集试验\color1-结合方法\seg_result_gray_repaint',
                  # r'D:\map_translate\看看效果\0627两个重绘数据集试验\color1-纯p2pHD\fake_result']

    # real_paths = [r'/data1/liangshuaizhe/map_translate/p2pHD-change-seg/GAN/src/pix2pixHD/TWmap16_log_2.1_seg_as_input_repaint1/real_result',
    #               r'/data1/liangshuaizhe/map_translate/p2pHD-change-seg/GAN/src/pix2pixHD/changemodel_log_3.0_addgradloss_0.1_0/real_result',
    #               r'/data1/liangshuaizhe/map_translate/p2pHD-change-seg/GAN/src/pix2pixHD/changemodel_log_3.1_addgradloss_0.2_0/real_result',
    #               r'/data1/liangshuaizhe/map_translate/p2pHD-change-seg/GAN/src/pix2pixHD/changemodel_log_3.4_addgradloss_0.5_0/real_result',
    #               r'/data1/liangshuaizhe/map_translate/p2pHD-change-seg/GAN/src/pix2pixHD/changemodel_log_3.2_addgradloss_0_0.1/real_result',
    #               r'/data1/liangshuaizhe/map_translate/p2pHD-change-seg/GAN/src/pix2pixHD/changemodel_log_3.3_addgradloss_0_0.2/real_result',
    #               r'/data1/liangshuaizhe/map_translate/p2pHD-change-seg/GAN/src/pix2pixHD/changemodel_log_3.5_addgradloss_0_0.5/real_result',
    #               r'/data1/liangshuaizhe/map_translate/p2pHD-change-seg/GAN/src/pix2pixHD/changemodel_log_4.0_connect_softmax/real_result',
    #               r'/data1/liangshuaizhe/map_translate/p2pHD-change-seg/GAN/src/pix2pixHD/changemodel_log_4.1_connect_featuremap/real_result',
    #               r'/data1/liangshuaizhe/map_translate/p2pHD-change-seg/GAN/src/pix2pixHD/changemodel_log_5.0_changedlv3p_upsample2conv/real_result',
    #               r'/data1/liangshuaizhe/map_translate/p2pHD-change-seg/GAN/src/pix2pixHD/changemodel_log_5.1_changedlv3p_clsconv_1x/real_result']
    #
    # fake_paths = []
    # for tmp in real_paths:
    #     fake_paths.append(tmp[:-11]+'fake_result')

    for i in range(len(real_paths)):
        real_path=real_paths[i]
        fake_path=fake_paths[i]
        real_files=sorted(make_dataset(real_path))
        fake_files = sorted(make_dataset(fake_path))
        assert len(real_files)==len(fake_files)
        ssim,mse, psnr,ssim_rgb = scores_from_floder(real_path, fake_path)
        sh.cell(row=i + 2, column=1, value=real_path)
        sh.cell(row=i+2, column=2, value=mse)
        sh.cell(row=i + 2, column=3, value=ssim)
        sh.cell(row=i + 2, column=4, value=psnr)
        sh.cell(row=i + 2, column=5, value=ssim_rgb)
        print(f'mse:{mse},ssim:{ssim},psnr:{psnr},ssim_rgb:{ssim_rgb}')

    wb.save('./data1.xlsx')